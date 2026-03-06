#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Amazon Sales Report Analyzer - 通用型销售报告分析工具
支持自动识别数据范围和可用国家，生成定制化报告

代码审计硬性要求 (必须遵守):
- 安全: 禁止禁用SSL验证, 禁止爬取敏感数据
- 健壮: 除零检查, 数组边界检查, 禁止裸except
- 验证: 数字解析必须处理异常, HTML必须escape用户数据
"""

import os
import re
import csv
import json
import argparse
import logging
from datetime import datetime
from collections import defaultdict
from html import escape
from typing import Dict, List, Any, Optional

# ==================== 配置加载 ====================
CONFIG_FILE = "config.json"
_config = None


def load_config() -> Dict[str, Any]:
    """加载配置文件"""
    global _config
    if _config is not None:
        return _config

    default_config = {
        "countries": {
            'DE': {'name': '德国', 'currency': 'EUR', 'symbol': '€'},
            'IT': {'name': '意大利', 'currency': 'EUR', 'symbol': '€'},
            'FR': {'name': '法国', 'currency': 'EUR', 'symbol': '€'},
            'ES': {'name': '西班牙', 'currency': 'EUR', 'symbol': '€'},
            'UK': {'name': '英国', 'currency': 'GBP', 'symbol': '£'},
            'US': {'name': '美国', 'currency': 'USD', 'symbol': '$'}
        },
        "exchange_rates": {"EUR_USD": 1.08, "GBP_USD": 1.27},
        "report": {
            "default_reports_dir": "amazon_reports",
            "sku_top_n": 50,
            "returns_top_n": 20,
            "comments_top_n": 15
        },
        "validation": {
            "required_fields": {
                "business": ["sku", "quantity", "sales"],
                "transaction": ["sku", "quantity", "sales"],
                "returns": ["sku", "reason"],
                "ads": ["country", "spend", "sales"]
            },
            "min_rows": 1,
            "warn_threshold": 1000
        }
    }

    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), CONFIG_FILE)
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                user_config = json.load(f)
                # 合并配置
            _config = {**default_config, **user_config}
            logging.info(f"已加载配置文件: {config_path}")
        except Exception as e:
            logging.warning(f"配置文件加载失败，使用默认配置: {e}")
            _config = default_config
    else:
        logging.info("未找到配置文件，使用默认配置")
        _config = default_config

    return _config


def get_country_map() -> Dict[str, Dict[str, str]]:
    """获取国家映射配置"""
    config = load_config()
    return config.get("countries", {})


def get_exchange_rates() -> Dict[str, float]:
    """获取汇率配置"""
    config = load_config()
    return config.get("exchange_rates", {"EUR_USD": 1.08, "GBP_USD": 1.27})


def get_report_config() -> Dict[str, Any]:
    """获取报告配置"""
    config = load_config()
    return config.get("report", {})


def get_validation_config() -> Dict[str, Any]:
    """获取数据验证配置"""
    config = load_config()
    return config.get("validation", {})


# ==================== 日志系统 ====================
def setup_logging(log_level: str = "INFO", log_file: str = None):
    """初始化日志系统"""
    level = getattr(logging, log_level.upper(), logging.INFO)

    # 创建格式化器
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # 获取根日志记录器
    logger = logging.getLogger()
    logger.setLevel(level)

    # 清除已有的处理器
    logger.handlers.clear()

    # 控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(level)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # 文件处理器(如果指定)
    if log_file:
        try:
            file_handler = logging.FileHandler(log_file, encoding='utf-8')
            file_handler.setLevel(level)
            file_handler.setFormatter(formatter)
            logger.addHandler(file_handler)
            logging.info(f"日志已写入文件: {log_file}")
        except Exception as e:
            logging.warning(f"无法创建日志文件: {e}")

    return logger



DEFAULT_REPORTS_DIR = "amazon_reports"
FALLBACK_RATES = {"EUR_USD": 1.08, "GBP_USD": 1.27}

COUNTRY_MAP = {
    'DE': {'name': '德国', 'currency': 'EUR', 'symbol': '€'},
    'IT': {'name': '意大利', 'currency': 'EUR', 'symbol': '€'},
    'FR': {'name': '法国', 'currency': 'EUR', 'symbol': '€'},
    'ES': {'name': '西班牙', 'currency': 'EUR', 'symbol': '€'},
    'UK': {'name': '英国', 'currency': 'GBP', 'symbol': '£'},
    'US': {'name': '美国', 'currency': 'USD', 'symbol': '$'}
}

# ==================== 工具函数 ====================
def sanitize_filename(name: str) -> str:
    """生成安全的文件名，移除不允许的字符"""
    if not name:
        return "report"
    # 移除或替换不允许在文件名中使用的字符
    import re
    # 替换空格和常见分隔符
    name = re.sub(r'[\s\-/:\\]', '_', name)
    # 移除非ASCII字符（保留中文但移除特殊符号）
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    # 移除连续的下划线
    while '__' in name:
        name = name.replace('__', '_')
    return name.strip('_')


# ==================== 汇率获取 ====================
def fetch_exchange_rates() -> Dict[str, float]:
    """获取汇率（优先使用配置文件）"""
    rates = get_exchange_rates()
    logging.info(f"汇率: EUR→USD: {rates['EUR_USD']:.4f}, GBP→USD: {rates['GBP_USD']:.4f}")
    print(f"汇率: EUR→USD: {rates['EUR_USD']:.4f}, GBP→USD: {rates['GBP_USD']:.4f}")
    return rates.copy()

# ==================== 工具函数 ====================
def parse_number(value: Any) -> float:
    """解析各种格式数字"""
    if not value or value == "":
        return 0.0
    value = str(value).strip()
    if '<' in value:
        return 0.0
    if '%' in value:
        value = value.replace('%', '')
        try:
            return float(value)
        except ValueError:
            return 0.0
    value = re.sub(r'[€£$US\s]', '', value)
    if ',' in value and '.' in value:
        if value.find(',') < value.find('.'):
            value = value.replace(',', '')
        else:
            value = value.replace('.', '').replace(',', '.')
    elif ',' in value:
        if len(value.split(',')[-1]) == 2:
            value = value.replace(',', '.')
        else:
            value = value.replace(',', '')
    try:
        return float(value) if value else 0.0
    except ValueError:
        return 0.0


# ==================== 数据验证 ====================
class ValidationResult:
    """验证结果类"""
    def __init__(self):
        self.warnings: List[str] = []
        self.errors: List[str] = []
        self.info: List[str] = []

    def add_warning(self, msg: str):
        self.warnings.append(msg)
        logging.warning(msg)

    def add_error(self, msg: str):
        self.errors.append(msg)
        logging.error(msg)

    def add_info(self, msg: str):
        self.info.append(msg)
        logging.info(msg)

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0

    def summary(self) -> str:
        lines = []
        if self.info:
            lines.append(f"信息: {len(self.info)}条")
        if self.warnings:
            lines.append(f"警告: {len(self.warnings)}条")
        if self.errors:
            lines.append(f"错误: {len(self.errors)}条")
        return ", ".join(lines) if lines else "验证通过"


def validate_csv_file(filepath: str, report_type: str) -> ValidationResult:
    """验证CSV文件的数据完整性"""
    result = ValidationResult()
    validation_config = get_validation_config()
    required_fields = validation_config.get("required_fields", {}).get(report_type, [])
    min_rows = validation_config.get("min_rows", 1)
    warn_threshold = validation_config.get("warn_threshold", 1000)

    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)

            # 读取表头
            try:
                headers = next(reader)
            except StopIteration:
                result.add_error(f"文件为空: {os.path.basename(filepath)}")
                return result

            headers_lower = [h.lower().strip() for h in headers]

            # 检查必需字段
            missing_fields = []
            for field in required_fields:
                found = False
                for h in headers_lower:
                    if field in h or h == field:
                        found = True
                        break
                if not found:
                    missing_fields.append(field)

            if missing_fields:
                result.add_warning(f"缺少字段 {missing_fields}: {os.path.basename(filepath)}")

            # 统计行数
            row_count = 0
            empty_rows = 0
            for row in reader:
                row_count += 1
                if not any(cell.strip() for cell in row):
                    empty_rows += 1

            # 检查行数
            if row_count < min_rows:
                result.add_warning(f"数据行数过少 ({row_count}): {os.path.basename(filepath)}")
            elif row_count > warn_threshold:
                result.add_info(f"大数据文件 ({row_count}行): {os.path.basename(filepath)}")

            # 检查空行
            if empty_rows > 0:
                result.add_warning(f"包含 {empty_rows} 个空行: {os.path.basename(filepath)}")

            if result.is_valid and not result.warnings:
                result.add_info(f"验证通过: {os.path.basename(filepath)}")

    except Exception as e:
        result.add_error(f"读取文件失败: {os.path.basename(filepath)}, 错误: {e}")

    return result


def validate_reports_directory(reports_dir: str) -> ValidationResult:
    """验证整个报告目录的数据完整性"""
    result = ValidationResult()
    result.add_info(f"开始验证目录: {reports_dir}")

    if not os.path.exists(reports_dir):
        result.add_error(f"目录不存在: {reports_dir}")
        return result

    files = os.listdir(reports_dir)
    csv_files = [f for f in files if f.endswith('.csv')]

    if not csv_files:
        result.add_error("目录中没有CSV文件")
        return result

    result.add_info(f"发现 {len(csv_files)} 个CSV文件")

    # 分类验证
    validation_summary = {}

    for f in csv_files:
        filepath = os.path.join(reports_dir, f)
        f_lower = f.lower()

        # 确定报告类型
        report_type = "unknown"
        if 'business' in f_lower or '业务' in f:
            report_type = "business"
        elif 'transaction' in f_lower or '交易' in f:
            report_type = "transaction"
        elif 'return' in f_lower or '退货' in f:
            report_type = "returns"
        elif 'campaign' in f_lower or '广告' in f:
            report_type = "ads"

        if report_type != "unknown":
            vr = validate_csv_file(filepath, report_type)
            if report_type not in validation_summary:
                validation_summary[report_type] = {"valid": 0, "warnings": 0, "errors": 0}

            if vr.is_valid:
                validation_summary[report_type]["valid"] += 1
            validation_summary[report_type]["warnings"] += len(vr.warnings)
            validation_summary[report_type]["errors"] += len(vr.errors)

    # 输出汇总
    for rtype, counts in validation_summary.items():
        result.add_info(f"  {rtype}: {counts['valid']}个通过, {counts['warnings']}个警告, {counts['errors']}个错误")

    result.add_info(f"验证完成: {result.summary()}")
    return result


# ==================== 文件扫描 ====================
def scan_reports_directory(reports_dir):
    """扫描报告目录，识别可用的数据文件"""
    if not os.path.exists(reports_dir):
        print(f"警告: 目录不存在 {reports_dir}")
        return {'countries': [], 'period': 'Unknown', 'files': {}}
    
    files = os.listdir(reports_dir)
    available_countries = set()
    file_types = {'business': [], 'transaction': [], 'returns': [], 'ads': []}
    period = "Unknown"
    
    for f in files:
        if not f.endswith('.csv'):
            continue
        
        # 识别国家
        for country in COUNTRY_MAP.keys():
            if country in f:
                available_countries.add(country)
        
        # 识别文件类型
        f_lower = f.lower()
        if 'business' in f_lower or '业务' in f:
            file_types['business'].append(f)
        elif 'transaction' in f_lower or '交易' in f:
            file_types['transaction'].append(f)
        elif 'return' in f_lower or '退货' in f:
            file_types['returns'].append(f)
        elif 'campaign' in f_lower or '广告' in f:
            file_types['ads'].append(f)
    
    # 从文件名智能提取周期
    def detect_period_from_text(text):
        """从文本（文件名或目录名）检测周期"""
        month_map = {'Jan': '1', 'Feb': '2', 'Mar': '3', 'Apr': '4', 'May': '5', 'Jun': '6',
                     'Jul': '7', 'Aug': '8', 'Sep': '9', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
        
        # 查找年份
        year_match = re.search(r'(20\d{2})', text)
        if not year_match:
            return None
        year = year_match.group(1)
        
        # 查找季度 Q1-Q4
        q_match = re.search(r'Q([1-4])', text, re.IGNORECASE)
        if q_match:
            return f"{year}年Q{q_match.group(1)}"
        
        # 查找月份 Jan-Dec
        for m, num in month_map.items():
            if m in text:
                return f"{year}年{num}月"
        
        # 只有年份
        return f"{year}年"
    
    # 从目录名和文件名提取周期
    dir_name = os.path.basename(reports_dir)
    period = "Unknown"
    
    # 收集所有文本用于检测，按优先级排序（transaction > business > returns）
    all_texts = [dir_name]
    for ftype in ['transaction', 'business', 'returns', 'ads']:
        if ftype in file_types:
            all_texts.extend(file_types[ftype])
    
    # 检测周期，优先使用更具体的月份/季度
    best_period = None
    for text in all_texts:
        p = detect_period_from_text(text)
        if p:
            # 如果检测到具体月份，优先使用
            if '月' in p or 'Q' in p:
                period = p
                break
            # 否则保存第一个年份结果作为备选
            if not best_period:
                best_period = p
    
    if period == "Unknown" and best_period:
        period = best_period
    
    return {
        'countries': sorted(list(available_countries)),
        'period': period,
        'files': file_types
    }

# ==================== 数据读取 ====================
def load_transaction_report(reports_dir, country):
    """从Transaction Report提取SKU销售数据"""
    all_files = os.listdir(reports_dir)
    files = [f for f in all_files if 'Transaction' in f and f.startswith(country)]
    if not files:
        files = [f for f in all_files if 'Transaction' in f and country in f]
    if not files:
        files = [f for f in all_files if 'Transaction' in f]
    if not files:
        return [], {}
    
    filepath = os.path.join(reports_dir, files[0])
    
    products = defaultdict(lambda: {'title': '', 'quantity': 0, 'sales': 0})
    
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)

        try:
            headers = next(reader)
        except StopIteration:
            return [], {}

        # 跳过注释行（德语报告有多行描述）
        try:
            while headers and len(headers) == 1 and not headers[0].strip().startswith('"'):
                headers = next(reader)
            while headers and not any('sku' in h.lower() or 'typ' in h.lower() or 'type' in h.lower() for h in headers if h):
                headers = next(reader)
        except StopIteration:
            return [], {}
        
        indices = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            h_lower = h.lower().replace('"', '')
            if h_lower == 'sku':
                indices['sku'] = i
            elif 'description' in h_lower and 'product' not in h_lower:
                indices['title'] = i
            elif h_lower in ['quantity', 'menge', 'quantité', 'cantidad', 'quantità']:
                if 'qty' not in indices:
                    indices['qty'] = i
            elif h_lower == 'product sales' or h_lower in ['umsätze', 'ventes de produits', 'ventas de productos', 'vendite prodotti', 'vendite']:
                if 'sales' not in indices:
                    indices['sales'] = i
            elif h_lower in ['type', 'typ', 'tipo']:
                indices['type'] = i
        
        for row in reader:
            if len(row) < 5:
                continue
            
            tx_type = row[indices.get('type', 0)].lower() if indices.get('type') and len(row) > indices.get('type', 0) else ""
            order_keywords = ['order', 'bestellung', 'commande', 'pedido', 'ordine']
            if not any(kw in tx_type for kw in order_keywords):
                continue
            
            sku = row[indices.get('sku', 0)].strip() if indices.get('sku') and len(row) > indices.get('sku', 0) else ""
            if not sku:
                continue
            
            title = row[indices.get('title', 0)].strip() if indices.get('title') and len(row) > indices.get('title', 0) else ""
            qty = parse_number(row[indices.get('qty', 0)]) if indices.get('qty') and len(row) > indices.get('qty', 0) else 0
            sales = parse_number(row[indices.get('sales', 0)]) if indices.get('sales') and len(row) > indices.get('sales', 0) else 0
            
            if sales < 0:
                continue
            
            products[sku]['title'] = title
            products[sku]['quantity'] += qty
            products[sku]['sales'] += sales
    
    product_list = [{'sku': k, 'title': v['title'], 'quantity': v['quantity'], 'sales': v['sales'], 'conv_rate': 0, 'sessions': 0} 
                   for k, v in products.items()]
    
    total_orders = sum(p['quantity'] for p in product_list)
    total_sales = sum(p['sales'] for p in product_list)
    
    summary = {
        'sessions': 0,
        'conversion_rate': 0,
        'orders': total_orders,
        'total_sales': total_sales
    }
    
    return product_list, summary


def load_business_report(reports_dir, country):
    """读取Business Report，如果没有则尝试从Transaction Report加载"""
    all_files = os.listdir(reports_dir)
    files = [f for f in all_files if 'BusinessReport' in f and f.startswith(country)]
    if not files:
        files = [f for f in all_files if 'BusinessReport' in f and country in f]
    
    if not files:
        return load_transaction_report(reports_dir, country)
    
    # 使用对应国家的BusinessReport文件
    filepath = os.path.join(reports_dir, files[0])
    
    products = []
    total_sessions = 0
    total_conv_rate = 0
    product_count = 0
    
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            return load_transaction_report(reports_dir, country)

        indices = {}
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if 'sku' in h_lower and 'sku' not in indices:
                indices['sku'] = i
            elif '标题' in h or 'title' in h_lower:
                if 'title' not in indices:
                    indices['title'] = i
            elif '已订购商品数量' in h or h == 'Ordered' or (h_lower == 'ordered' and 'b2b' not in h_lower):
                if 'qty' not in indices and '- b2b' not in h_lower:
                    indices['qty'] = i
            elif '已订购商品销售额' in h or h == 'Sales' or (h_lower == 'sales' and 'b2b' not in h_lower and 'refund' not in h_lower):
                if 'sales' not in indices and '- b2b' not in h_lower:
                    indices['sales'] = i
            elif ('会话' in h or 'session' in h_lower) and 'sessions' not in indices:
                if '总计' in h or 'total' in h_lower:
                    indices['sessions'] = i
            elif ('商品会话百分比' in h or '转化率' in h or 'conversion' in h_lower) and 'conv_rate' not in indices:
                if 'b2b' not in h_lower:
                    indices['conv_rate'] = i

        for row in reader:
            if len(row) <= max(indices.get('sku', 0), indices.get('qty', 0), indices.get('sales', 0)):
                continue
            
            sku = row[indices.get('sku', 0)].strip() if indices.get('sku') else ""
            title = row[indices.get('title', 0)].strip() if indices.get('title') else ""
            qty = parse_number(row[indices.get('qty', 0)]) if indices.get('qty') else 0
            sales = parse_number(row[indices.get('sales', 0)]) if indices.get('sales') else 0
            
            sku_sessions = 0
            sku_conv_rate = 0
            if indices.get('sessions') and len(row) > indices['sessions']:
                sku_sessions = int(parse_number(row[indices['sessions']]))
                total_sessions += sku_sessions
            
            if indices.get('conv_rate') and len(row) > indices['conv_rate']:
                cr = parse_number(row[indices['conv_rate']])
                if cr > 0:
                    total_conv_rate += cr
                    product_count += 1
                    sku_conv_rate = cr
            
            if sku:
                products.append({'sku': sku, 'title': title, 'quantity': qty, 'sales': sales, 'conv_rate': sku_conv_rate, 'sessions': sku_sessions})
    
    avg_conv = total_conv_rate / product_count if product_count > 0 else 0
    
    summary = {
        'sessions': total_sessions,
        'conversion_rate': avg_conv,
        'orders': sum(p['quantity'] for p in products),
        'total_sales': sum(p['sales'] for p in products)
    }
    
    return products, summary

def load_returns_report(reports_dir):
    """读取Returns Report"""
    files = [f for f in os.listdir(reports_dir) if 'return' in f.lower() or '退货' in f]
    if not files:
        return [], {}, {}, []
    
    filepath = os.path.join(reports_dir, files[0])
    returns = []
    sku_by_warehouse = defaultdict(lambda: defaultdict(int))
    sku_with_comments = []
    
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            return [], {}, {}, []

        indices = {}
        for i, h in enumerate(headers):
            h_lower = h.lower().replace('"', '')
            if h_lower == 'sku':
                indices['sku'] = i
            elif 'product' in h_lower:
                indices['product'] = i
            elif 'reason' in h_lower:
                indices['reason'] = i
            elif 'fulfillment' in h_lower or 'warehouse' in h_lower or 'center' in h_lower:
                indices['warehouse'] = i
            elif 'customer' in h_lower and 'comment' in h_lower:
                indices['comments'] = i
        
        for row in reader:
            if len(row) > max(indices.get('sku', 0), indices.get('reason', 0)):
                sku = row[indices.get('sku', 0)].strip() if indices.get('sku') else ""
                reason = row[indices.get('reason', 0)].strip() if indices.get('reason') else ""
                warehouse = row[indices.get('warehouse', 0)].strip() if indices.get('warehouse') else "Unknown"
                comments = row[indices.get('comments', 0)].strip() if indices.get('comments') else ""
                
                returns.append({
                    'sku': sku,
                    'reason': reason,
                    'warehouse': warehouse,
                    'comments': comments
                })
                
                # 按仓库统计
                if sku:
                    sku_by_warehouse[sku][warehouse] += 1
                
                # 有客户评论的记录
                if comments and sku:
                    sku_with_comments.append({
                        'sku': sku,
                        'reason': reason,
                        'comments': comments
                    })
    
    return returns, {}, sku_by_warehouse, sku_with_comments

def load_ads_report(reports_dir):
    """读取广告报表"""
    files = [f for f in os.listdir(reports_dir) if 'campaign' in f.lower() or '广告' in f]
    if not files:
        return {}, {}
    
    filepath = os.path.join(reports_dir, files[0])
    ads_by_country = defaultdict(lambda: {'impressions': 0, 'clicks': 0, 'spend': 0, 'sales': 0, 'orders': 0})
    
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            return {}, {}

        indices = {}
        for i, h in enumerate(headers):
            h_cn = h.strip().replace('﻿', '')
            if '国家' in h_cn:
                indices['country'] = i
            elif '总成本' in h_cn and '转换' not in h_cn:
                indices['spend'] = i
            elif '销售额' in h_cn and '预算' not in h_cn and '转换' not in h_cn:
                indices['sales'] = i
            elif '点击量' in h_cn and 'clicks' not in indices:
                indices['clicks'] = i
            elif '展示量' in h_cn and 'impressions' not in indices:
                indices['impressions'] = i
            elif '购买量' in h_cn and 'orders' not in indices:
                indices['orders'] = i
        
        for row in reader:
            if len(row) < 5:
                continue
            
            country_raw = row[indices.get('country', 0)].strip() if indices.get('country') else ""
            country_map = {'德国': 'DE', '意大利': 'IT', '法国': 'FR', '西班牙': 'ES', '英国': 'UK'}
            country = country_map.get(country_raw, None)
            
            if country is None:
                continue
            
            spend = parse_number(row[indices.get('spend', 0)]) if indices.get('spend') else 0
            sales = parse_number(row[indices.get('sales', 0)]) if indices.get('sales') else 0
            clicks = parse_number(row[indices.get('clicks', 0)]) if indices.get('clicks') else 0
            impressions = parse_number(row[indices.get('impressions', 0)]) if indices.get('impressions') else 0
            orders = parse_number(row[indices.get('orders', 0)]) if indices.get('orders') else 0
            
            ads_by_country[country]['spend'] += spend
            ads_by_country[country]['sales'] += sales
            ads_by_country[country]['clicks'] += clicks
            ads_by_country[country]['impressions'] += impressions
            ads_by_country[country]['orders'] += orders
    
    return dict(ads_by_country), {}



# ==================== 数据导出 ====================
def export_to_csv(data, filepath):
    """导出数据到CSV文件"""
    import csv
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['=== 国家汇总 ==='])
        writer.writerow(['国家', '销售额(本地)', '销售额(USD)', '订单数', '客单价', '转化率', '广告花费', 'ROI', 'CTR', 'CVR', 'ACOS', '退货数'])
        for row in data.get('country_summary', []):
            writer.writerow(row)
        writer.writerow([])
        writer.writerow(['=== SKU明细 ==='])
        writer.writerow(['SKU', '国家', '数量', '销售额', '转化率', '会话数'])
        for row in data.get('sku_details', []):
            writer.writerow(row)
        writer.writerow([])
        writer.writerow(['=== 退货明细 ==='])
        writer.writerow(['SKU', '原因', '仓库', '数量'])
        for row in data.get('returns_details', []):
            writer.writerow(row)
        writer.writerow([])
        writer.writerow(['=== 广告明细 ==='])
        writer.writerow(['国家', '花费', '点击', '展示', '销售', '订单', 'CTR', 'CVR', 'ACOS'])
        for row in data.get('ads_details', []):
            writer.writerow(row)
    print(f"CSV导出完成: {filepath}")


def export_to_excel(data, filepath, rates=None):
    """导出数据到Excel文件（数值使用数字格式，USD销售额使用公式）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, numbers

    FORMAT_DECIMAL2 = '0.00'
    FORMAT_DECIMAL0 = '0'
    FORMAT_PERCENT = '0.00%'
    FORMAT_RAW = numbers.FORMAT_NUMBER

    if rates is None:
        rates = {'EUR_USD': 1.08, 'GBP_USD': 1.27}

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 国家汇总
    ws1 = wb.create_sheet("国家汇总")

    # 汇率设置区域
    ws1.merge_cells('A1:D1')
    ws1.cell(1, 1, "汇率设置").font = Font(bold=True, size=12)
    ws1.cell(1, 1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws1.cell(2, 1, "EUR→USD 汇率:")
    ws1.cell(2, 2, rates.get('EUR_USD', 1.08))
    ws1.cell(2, 2).number_format = '0.0000'

    ws1.cell(3, 1, "GBP→USD 汇率:")
    ws1.cell(3, 2, rates.get('GBP_USD', 1.27))
    ws1.cell(3, 2).number_format = '0.0000'

    # 表头
    headers1 = ['国家', '销售额(本地)', '销售额(USD)', '订单数', '客单价', '转化率', '广告花费', 'ROI', 'CTR', 'CVR', 'ACOS', '退货数']
    col_formats1 = [FORMAT_RAW, FORMAT_DECIMAL2, FORMAT_DECIMAL2, FORMAT_DECIMAL0,
                    FORMAT_DECIMAL2, FORMAT_PERCENT, FORMAT_DECIMAL2, FORMAT_DECIMAL2,
                    FORMAT_PERCENT, FORMAT_PERCENT, FORMAT_PERCENT, FORMAT_DECIMAL0]

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(5, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 数据
    for row_idx, row in enumerate(data.get('country_summary', []), 6):
        country_name = row[0]
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row_idx, col_idx, value)
            if col_idx == 3:  # USD销售额使用公式
                if '德国' in country_name or '西班牙' in country_name or '法国' in country_name or '意大利' in country_name:
                    cell.value = f'=B{row_idx}*$B$2'
                elif '英国' in country_name:
                    cell.value = f'=B{row_idx}*$B$3'
                else:
                    cell.value = f'=B{row_idx}'
                cell.number_format = FORMAT_DECIMAL2
            elif col_idx > 1 and col_formats1[col_idx - 1]:
                cell.number_format = col_formats1[col_idx - 1]

    # 列宽
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws1.column_dimensions[col].width = 10

    # Sheet 2: SKU明细
    ws2 = wb.create_sheet("SKU明细")
    headers2 = ['SKU', '国家', '数量', '销售额', '转化率', '会话数']
    col_formats2 = [FORMAT_RAW, FORMAT_RAW, FORMAT_DECIMAL0, FORMAT_DECIMAL2, FORMAT_PERCENT, FORMAT_DECIMAL0]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for row_idx, row in enumerate(data.get('sku_details', []), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row_idx, col_idx, value)
            if col_idx >= 3 and col_formats2[col_idx - 1]:
                cell.number_format = col_formats2[col_idx - 1]

    # Sheet 3: 退货明细
    ws3 = wb.create_sheet("退货明细")
    headers3 = ['SKU', '原因', '仓库', '数量']

    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for row_idx, row in enumerate(data.get('returns_details', []), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row_idx, col_idx, value)
            if col_idx == 4:
                cell.number_format = FORMAT_DECIMAL0

    # Sheet 4: 广告明细
    ws4 = wb.create_sheet("广告明细")
    headers4 = ['国家', '花费', '点击', '展示', '销售', '订单', 'CTR', 'CVR', 'ACOS']
    col_formats4 = [FORMAT_RAW, FORMAT_DECIMAL2, FORMAT_DECIMAL0, FORMAT_DECIMAL0,
                    FORMAT_DECIMAL2, FORMAT_DECIMAL0, FORMAT_PERCENT, FORMAT_PERCENT, FORMAT_PERCENT]

    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for row_idx, row in enumerate(data.get('ads_details', []), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws4.cell(row_idx, col_idx, value)
            if col_idx > 1 and col_formats4[col_idx - 1]:
                cell.number_format = col_formats4[col_idx - 1]

    wb.save(filepath)
    print(f"Excel导出完成: {filepath}")


def prepare_export_data(country_data, ads_data, returns, rates):
    """准备导出数据"""
    from collections import defaultdict
    export_data = {
        'country_summary': [],
        'sku_details': [],
        'returns_details': [],
        'ads_details': []
    }

    country_returns = defaultdict(int)
    for r in returns:
        country_returns[r.get('country', '')] += 1

    for country, d in country_data.items():
        cfg = COUNTRY_MAP.get(country, {})
        currency = cfg.get('currency', 'USD')

        sales = d.get('total_sales', 0)
        orders = int(d.get('orders', 0))
        sessions = int(d.get('sessions', 0))
        conversion_rate = d.get('conversion_rate', 0) / 100

        aov = sales / orders if orders > 0 else 0
        sales_usd = sales * rates['EUR_USD'] if currency == 'EUR' else sales * rates['GBP_USD'] if currency == 'GBP' else sales

        ads = ads_data.get(country, {})
        ads_spend = ads.get('spend', 0)
        impressions = ads.get('impressions', 0)
        clicks = ads.get('clicks', 0)
        ads_sales = ads.get('sales', 0)
        ads_orders = ads.get('orders', 0)

        roi = ads_sales / ads_spend if ads_spend > 0 else 0
        ctr = clicks / impressions if impressions > 0 else 0
        cvr = ads_orders / clicks if clicks > 0 else 0
        acos = ads_spend / ads_sales if ads_sales > 0 else 0

        returns_count = country_returns.get(country, 0)

        export_data['country_summary'].append([
            cfg.get('name', country),
            sales,
            sales_usd,
            orders,
            aov,
            conversion_rate if conversion_rate > 0 else '',
            ads_spend,
            roi if roi > 0 else '',
            ctr if ctr > 0 else '',
            cvr if cvr > 0 else '',
            acos if acos > 0 else '',
            returns_count
        ])

        for product in d.get('products', []):
            sku_conv_rate = product.get('conv_rate', 0) / 100 if product.get('conv_rate', 0) > 0 else 0
            sku_sessions = int(product.get('sessions', 0))
            export_data['sku_details'].append([
                product.get('sku', ''),
                cfg.get('name', country),
                product.get('quantity', 0),
                product.get('sales', 0),
                sku_conv_rate if sku_conv_rate > 0 else '',
                sku_sessions
            ])

    for r in returns:
        export_data['returns_details'].append([
            r.get('sku', ''),
            r.get('reason', ''),
            r.get('warehouse', ''),
            1
        ])

    for country, ads in ads_data.items():
        cfg = COUNTRY_MAP.get(country, {})
        spend = ads.get('spend', 0)
        clicks = int(ads.get('clicks', 0))
        impressions = int(ads.get('impressions', 0))
        sales = ads.get('sales', 0)
        orders = int(ads.get('orders', 0))

        ctr = clicks / impressions if impressions > 0 else 0
        cvr = orders / clicks if clicks > 0 else 0
        acos = spend / sales if sales > 0 else 0

        export_data['ads_details'].append([
            cfg.get('name', country),
            spend,
            clicks,
            impressions,
            sales,
            orders,
            ctr if ctr > 0 else '',
            cvr if cvr > 0 else '',
            acos if acos > 0 else ''
        ])

    return export_data

# ==================== 报告生成 ====================
def generate_report(reports_dir, output_file=None, countries=None, period=None, export_format=None, export_filename=None, eur_rate=None, gbp_rate=None):
    """生成HTML报告"""
    
    # 自动扫描
    scan = scan_reports_directory(reports_dir)
    available_countries = scan['countries']
    
    # 确定周期
    if not period:
        period = scan['period']

    # 转义period防止XSS
    period_escaped = escape(period) if period else 'Unknown'
    
    # 确定要分析的国家
    if countries:
        target_countries = [c for c in countries if c in available_countries]
    else:
        target_countries = available_countries
    
    if not target_countries:
        print("未找到可用的国家数据")
        return None
    
    # 汇率
    print("\n正在获取汇率...")
    # 汇率
    if eur_rate is not None or gbp_rate is not None:
        # 使用命令行指定的汇率
        default_rates = fetch_exchange_rates()
        rates = {
            'EUR_USD': eur_rate if eur_rate is not None else default_rates.get('EUR_USD', 1.08),
            'GBP_USD': gbp_rate if gbp_rate is not None else default_rates.get('GBP_USD', 1.27)
        }
        print(f"汇率 (自定义): EUR→USD: {rates['EUR_USD']:.4f}, GBP→USD: {rates['GBP_USD']:.4f}")
    else:
        rates = fetch_exchange_rates()
    
    # 加载各国数据
    country_data = {}
    has_business_data = False
    has_transaction_data = False
    
    for country in target_countries:
        products, summary = load_business_report(reports_dir, country)
        has_sessions = summary.get('sessions', 0) > 0
        has_conv = summary.get('conversion_rate', 0) > 0
        if has_sessions or has_conv:
            has_business_data = True
        if summary.get('total_sales', 0) > 0 or summary.get('orders', 0) > 0:
            has_transaction_data = True
        country_data[country] = {
            'products': products,
            'sessions': summary.get('sessions', 0),
            'conversion_rate': summary.get('conversion_rate', 0),
            'orders': summary.get('orders', 0),
            'total_sales': summary.get('total_sales', 0),
            'has_business': has_sessions or has_conv
        }
    
    # 加载退货数据
    returns, _, sku_by_warehouse, sku_with_comments = load_returns_report(reports_dir)
    sku_returns = defaultdict(lambda: {'count': 0, 'reasons': defaultdict(int)})
    reason_dist = defaultdict(int)
    total_returns = len(returns)
    has_returns = total_returns > 0
    for r in returns:
        sku_returns[r['sku']]['count'] += 1
        if r['reason']:
            sku_returns[r['sku']]['reasons'][r['reason']] += 1
            reason_dist[r['reason']] += 1
    
    top_returns = []
    for sku, data in sku_returns.items():
        main_reason = ''
        if data['reasons']:
            try:
                main_reason = max(data['reasons'].items(), key=lambda x: x[1])[0]
            except (ValueError, KeyError):
                main_reason = ''
        top_returns.append({'sku': sku, 'count': data['count'], 'main_reason': main_reason})
    top_returns = sorted(top_returns, key=lambda x: x['count'], reverse=True)[:10]
    
    # 加载广告数据
    ads_data, _ = load_ads_report(reports_dir)
    has_ads = any(d.get('spend', 0) > 0 for d in ads_data.values())
    
    # 计算汇总
    total_sessions = sum(d['sessions'] for d in country_data.values())
    total_orders = sum(d['orders'] for d in country_data.values())
    total_sales_eur = sum(d['total_sales'] for c, d in country_data.items() if COUNTRY_MAP[c]['currency'] == 'EUR')
    total_sales_gbp = sum(d['total_sales'] for c, d in country_data.items() if COUNTRY_MAP[c]['currency'] == 'GBP')
    total_sales_usd = total_sales_eur * rates['EUR_USD'] + total_sales_gbp * rates['GBP_USD']
    
    # 生成表格
    rows = ""
    global_sessions = 0
    global_orders = 0
    global_sales = 0
    global_ads_spend = 0
    
    for country in target_countries:
        d = country_data[country]
        cfg = COUNTRY_MAP[country]
        
        sessions = d['sessions']
        orders = d['orders']
        sales = d['total_sales']
        ads = ads_data.get(country, {})
        ads_spend = ads.get('spend', 0)
        
        aov = sales / orders if orders > 0 else 0
        roi = sales / ads_spend if ads_spend > 0 else 0
        
        sales_usd = sales * rates['EUR_USD'] if cfg['currency'] == 'EUR' else sales * rates['GBP_USD']
        ads_spend_usd = ads_spend * rates['EUR_USD'] if cfg['currency'] == 'EUR' else ads_spend * rates['GBP_USD']
        
        global_sessions += sessions
        global_orders += orders
        global_sales += sales_usd
        global_ads_spend += ads_spend_usd
        
        if has_ads:
            ctr = (ads.get('clicks', 0) / ads.get('impressions', 1) * 100) if ads.get('impressions', 0) > 0 else 0
            cvr = (ads.get('orders', 0) / ads.get('clicks', 1) * 100) if ads.get('clicks', 0) > 0 else 0
            acos = (ads_spend / ads.get('sales', 1) * 100) if ads.get('sales', 0) > 0 else 0
            session_cell = f'<td class="num">{int(sessions):,}</td>' if has_business_data else ''
            conv_cell = f'<td class="num">{d["conversion_rate"]:.1f}%</td>' if has_business_data else ''
            rows += f"""<tr>
                <td>{cfg['name']}</td>
                {session_cell}
                <td class="num">{cfg['symbol']}{sales:,.2f}</td>
                <td class="num">${sales_usd:,.2f}</td>
                <td class="num">{int(orders)}</td>
                <td class="num">{cfg['symbol']}{aov:,.2f}</td>
                {conv_cell}
                <td class="num">{cfg['symbol']}{ads_spend:,.2f} (${ads_spend_usd:,.2f})</td>
                <td class="num">{roi:.2f}</td>
                <td class="num">{ctr:.2f}%</td>
                <td class="num">{cvr:.2f}%</td>
                <td class="num">{acos:.2f}%</td>
            </tr>"""
        else:
            session_cell = f'<td class="num">{int(sessions):,}</td>' if has_business_data else ''
            conv_cell = f'<td class="num">{d["conversion_rate"]:.1f}%</td>' if has_business_data else ''
            rows += f"""<tr>
                <td>{cfg['name']}</td>
                {session_cell}
                <td class="num">{cfg['symbol']}{sales:,.2f}</td>
                <td class="num">${sales_usd:,.2f}</td>
                <td class="num">{int(orders)}</td>
                <td class="num">{cfg['symbol']}{aov:,.2f}</td>
                {conv_cell}
            </tr>"""
    
    # 全球总计
    global_aov = global_sales / global_orders if global_orders > 0 else 0
    global_roi = global_sales / global_ads_spend if global_ads_spend > 0 else 0
    
    if has_ads:
        global_ctr = 0
        global_cvr = 0
        total_clicks = sum(d.get('clicks', 0) for d in ads_data.values())
        total_impressions = sum(d.get('impressions', 0) for d in ads_data.values())
        if total_impressions > 0:
            global_ctr = total_clicks / total_impressions * 100
        if total_clicks > 0:
            global_cvr = sum(d.get('orders', 0) for d in ads_data.values()) / total_clicks * 100
        else:
            global_cvr = 0
        
        session_cell = f'<td class="num">{int(global_sessions):,}</td>' if has_business_data else ''
        conv_cell = '<td class="num">-</td>' if has_business_data else ''
        rows += f"""<tr style="font-weight:bold;background:#f0f4ff;">
            <td>全球总计</td>
            {session_cell}
            <td class="num">-</td>
            <td class="num">${global_sales:,.2f}</td>
            <td class="num">{int(global_orders)}</td>
            <td class="num">${global_aov:,.2f}</td>
            {conv_cell}
            <td class="num">${global_ads_spend:,.2f}</td>
            <td class="num">{global_roi:.2f}</td>
            <td class="num">{global_ctr:.2f}%</td>
            <td class="num">{global_cvr:.2f}%</td>
            <td class="num">-</td>
        </tr>"""
    else:
        session_cell = f'<td class="num">{int(global_sessions):,}</td>' if has_business_data else ''
        conv_cell = '<td class="num">-</td>' if has_business_data else ''
        rows += f"""<tr style="font-weight:bold;background:#f0f4ff;">
            <td>全球总计</td>
            {session_cell}
            <td class="num">-</td>
            <td class="num">${global_sales:,.2f}</td>
            <td class="num">{int(global_orders)}</td>
            <td class="num">${global_aov:,.2f}</td>
            {conv_cell}
        </tr>"""
    
    # SKU详细数据列表
    sku_rows = ""
    all_skus = []
    for country in target_countries:
        cfg = COUNTRY_MAP[country]
        for p in country_data[country]['products']:
            if p['sales'] > 0 or p['quantity'] > 0:
                sales_usd = p['sales'] * rates['EUR_USD'] if cfg['currency'] == 'EUR' else p['sales'] * rates['GBP_USD']
                all_skus.append({
                    'sku': p['sku'],
                    'title': p['title'],
                    'country': cfg['name'],
                    'currency': cfg['currency'],
                    'quantity': p['quantity'],
                    'sales': p['sales'],
                    'sales_usd': sales_usd,
                    'conv_rate': p.get('conv_rate', 0),
                    'sessions': p.get('sessions', 0)
                })
    
    all_skus = sorted(all_skus, key=lambda x: x['sales_usd'], reverse=True)
    total_sku_sessions = sum(s.get('sessions', 0) for s in all_skus)
    total_sku_qty = sum(s.get('quantity', 0) for s in all_skus)
    weighted_conv_sum = sum(s.get('sessions', 0) * s.get('conv_rate', 0) for s in all_skus)
    avg_conv_rate = weighted_conv_sum / total_sku_sessions if total_sku_sessions > 0 else 0
    
    # 获取国家货币符号
    country_symbols = {COUNTRY_MAP[c]['name']: COUNTRY_MAP[c]['symbol'] for c in target_countries}
    
    for item in all_skus[:50]:
        conv_rate = item.get('conv_rate', 0)
        conv_rate_str = f"{conv_rate:.2f}%" if conv_rate > 0 else "-"
        symbol = country_symbols.get(item['country'], '€')
        
        session_cell = f'<td class="num">{item.get("sessions", 0):,}</td>' if has_business_data else ''
        conv_cell = f'<td class="num">{conv_rate_str}</td>' if has_business_data else ''
        has_sessions_attr = 'true' if has_business_data else 'false'
        
        sku_rows += f"""<tr data-currency="{item['currency']}" data-has-sessions="{has_sessions_attr}">
            <td style="min-width:180px"><code>{escape(item['sku'])}</code></td>
            {session_cell}
            <td>{item['country']}</td>
            <td class="num">{int(item['quantity'])}</td>
            {conv_cell}
            <td class="num">{symbol}{item['sales']:,.2f}</td>
            <td class="num" data-usd="{item['sales_usd']:.2f}">${item['sales_usd']:,.2f}</td>
        </tr>"""

    # 准备图表数据
    # 1. 各国家销售额数据
    sales_chart_data = []
    orders_chart_data = []
    country_names_list = []
    for country in target_countries:
        d = country_data[country]
        cfg = COUNTRY_MAP[country]
        sales_usd = d['total_sales'] * rates['EUR_USD'] if cfg['currency'] == 'EUR' else d['total_sales'] * rates['GBP_USD']
        sales_chart_data.append(f"{{name: '{cfg['name']}', value: {sales_usd:.2f}}}")
        orders_chart_data.append(f"{{name: '{cfg['name']}', value: {d['orders']}}}")
        country_names_list.append(f"'{cfg['name']}'")

    sales_chart_data_str = "[" + ", ".join(sales_chart_data) + "]" if sales_chart_data else "[]"
    orders_chart_data_str = "[" + ", ".join(orders_chart_data) + "]" if orders_chart_data else "[]"
    country_names_list_str = "[" + ", ".join(country_names_list) + "]" if country_names_list else "[]"

    # 2. Top 10 SKU 销售额
    top_10_skus = all_skus[:10]
    sku_chart_data = []
    for item in top_10_skus:
        sku_name = item['sku']  # 显示完整SKU名称
        sku_chart_data.append(f"{{name: '{escape(sku_name)}', value: {item['sales_usd']:.2f}}}")
    sku_chart_data_str = "[" + ", ".join(sku_chart_data) + "]" if sku_chart_data else "[]"

    # 3. 退货原因数据
    returns_reason_data = []
    for reason, count in sorted(reason_dist.items(), key=lambda x: x[1], reverse=True)[:8]:
        returns_reason_data.append(f"{{name: '{escape(reason)}', value: {count}}}")
    returns_reason_data_str = "[" + ", ".join(returns_reason_data) + "]" if returns_reason_data else "[]"

    # 退货表格 - 按仓库分组
    returns_rows = ""
    for sku, warehouses in sorted(sku_by_warehouse.items(), key=lambda x: sum(x[1].values()), reverse=True)[:20]:
        total_count = sum(warehouses.values())
        warehouse_str = ", ".join([f"{w}: {c}" for w, c in warehouses.items()])
        returns_rows += f"<tr><td><code>{escape(sku)}</code></td><td class=\"num\">{total_count}</td><td style=\"font-size:11px;color:#666\">{escape(warehouse_str)}</td></tr>"
    
    # 退货原因统计
    reason_rows = ""
    for reason, count in sorted(reason_dist.items(), key=lambda x: x[1], reverse=True):
        pct = count / total_returns * 100 if total_returns > 0 else 0
        reason_rows += f"<tr><td>{escape(reason)}</td><td class=\"num\">{count}</td><td class=\"num\">{pct:.0f}%</td></tr>"
    
    # 客户评论
    comments_rows = ""
    for item in sku_with_comments[:15]:
        comments_rows += f"<tr><td><code>{escape(item['sku'])}</code></td><td>{escape(item['reason'])}</td><td style=\"text-align:left;white-space:pre-wrap;word-break:break-word\">{escape(item['comments'])}</td></tr>"
    
    # 退货模块HTML（无退货数据时隐藏）
    if has_returns:
        returns_html = f"""
        <div class="grid-2">
            <div class="card">
                <div class="card-title">🔄 退货产品</div>
                <table id="returnsTable">
                    <thead><tr><th>SKU</th><th class="num">数量</th><th>仓库</th></tr></thead>
                    <tbody>{returns_rows}</tbody>
                    <tfoot>
                        <tr style="font-weight:bold;background:#f0f4ff;">
                            <td>合计</td>
                            <td class="num" id="returnsQtySum">0</td>
                            <td></td>
                        </tr>
                    </tfoot>
                </table>
            </div>
            <div class="card">
                <div class="card-title">📈 退货原因</div>
                <table id="reasonTable">
                    <thead><tr><th>原因</th><th class="num">数量</th><th class="num">占比</th></tr></thead>
                    <tbody>{reason_rows}</tbody>
                    <tfoot>
                        <tr style="font-weight:bold;background:#f0f4ff;">
                            <td>合计</td>
                            <td class="num" id="reasonQtySum">0</td>
                            <td class="num">100%</td>
                        </tr>
                    </tfoot>
                </table>
            </div>
        </div>
        
        <div class="card">
            <div class="card-title">💬 客户评论 (有评论的退货)</div>
            <table>
                <thead><tr><th style="text-align:left">SKU</th><th style="text-align:left">原因</th><th style="text-align:left">客户评论</th></tr></thead>
                <tbody>{comments_rows}</tbody>
            </table>
        </div>
        """
    else:
        returns_html = ""
    
    # SKU筛选选项
    country_options = ''.join([f'<option value="{COUNTRY_MAP[c]["name"]}">{COUNTRY_MAP[c]["name"]}</option>' for c in target_countries])
    
    # 汇率选项
    eur_rate = rates['EUR_USD']
    gbp_rate = rates['GBP_USD']
    
    # HTML表头 - 根据数据来源动态生成
    if has_business_data:
        session_header = "<th>流量</th>"
        conv_header = "<th>转化率</th>"
    else:
        session_header = ""
        conv_header = ""
    
    if has_ads:
        core_headers = f"<th>国家</th>{session_header}<th>销售额(本地)</th><th>销售额(USD)</th><th>订单</th><th>客单价</th>{conv_header}<th>广告费</th><th>ROI</th><th>CTR</th><th>转化</th><th>ACOS</th>"
    else:
        core_headers = f"<th>国家</th>{session_header}<th>销售额(本地)</th><th>销售额(USD)</th><th>订单</th><th>客单价</th>{conv_header}"
    
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>亚马逊销售报告 - {period_escaped}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f7fa; padding: 16px; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        h1 {{ color: #1a1a2e; text-align: center; margin-bottom: 8px; font-size: 24px; }}
        .subtitle {{ text-align: center; color: #666; margin-bottom: 20px; font-size: 13px; }}
        .card {{ background: white; border-radius: 10px; padding: 16px; margin-bottom: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        .card-title {{ font-size: 15px; font-weight: 600; color: #1a1a2e; margin-bottom: 12px; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
        th, td {{ padding: 8px 6px; text-align: right; border-bottom: 1px solid #eee; }}
        th {{ background: #f8f9fc; font-weight: 600; color: #555; cursor: pointer; }}
        th:hover {{ background: #e8e9fc; }}
        th.sort-asc::after {{ content: " ▲"; font-size: 10px; }}
        th.sort-desc::after {{ content: " ▼"; font-size: 10px; }}
        td.num {{ font-variant-numeric: tabular-nums; }}
        code {{ background: #f5f5f5; padding: 2px 4px; border-radius: 3px; }}
        .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        .filter-bar {{ margin-bottom: 10px; display: flex; gap: 10px; align-items: center; }}
        .filter-bar label {{ font-size: 12px; color: #666; }}
        .filter-bar input, .filter-bar select {{ padding: 4px 8px; border: 1px solid #ddd; border-radius: 4px; font-size: 12px; }}
        .filter-bar input[type="text"] {{ width: 150px; }}
        .rate-bar {{ display: flex; justify-content: center; gap: 20px; margin-bottom: 16px; align-items: center; }}
        .rate-bar label {{ font-size: 12px; color: #666; }}
        .rate-bar input {{ width: 140px; padding: 4px 8px; border: 1px solid #ddd; border-radius: 4px; font-size: 12px; text-align: right; }}
        .rate-bar button {{ padding: 4px 12px; background: #4CAF50; color: white; border: none; border-radius: 4px; font-size: 12px; cursor: pointer; }}
        .rate-bar button:hover {{ background: #45a049; }}
        .chart-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 16px; }}
        .chart-card {{ background: white; border-radius: 10px; padding: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        .chart-card h3 {{ font-size: 14px; font-weight: 600; color: #1a1a2e; margin-bottom: 12px; }}
        .chart {{ width: 100%; height: 280px; }}
        .summary-cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 16px; }}
        .summary-card {{ background: white; border-radius: 10px; padding: 16px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        .summary-card .label {{ font-size: 12px; color: #666; margin-bottom: 4px; }}
        .summary-card .value {{ font-size: 20px; font-weight: 600; color: #1a1a2e; }}
        .summary-card .value.currency {{ color: #4CAF50; }}
        .summary-card .value.orders {{ color: #2196F3; }}
        .summary-card .value.returns {{ color: #f44336; }}
        .summary-card .value.ads {{ color: #FF9800; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 亚马逊销售报告 - {period_escaped}</h1>
        <p class="subtitle">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <div class="rate-bar">
            <label>汇率 (EUR→USD):</label>
            <input type="number" id="rateEurUsd" value="{eur_rate:.4f}" step="0.0001">
            <label>汇率 (GBP→USD):</label>
            <input type="number" id="rateGbpUsd" value="{gbp_rate:.4f}" step="0.0001">
            <button onclick="recalcRates()">重新计算</button>
        </div>

        <!-- 汇总卡片 -->
        <div class="summary-cards">
            <div class="summary-card">
                <div class="label">总销售额</div>
                <div class="value currency">${global_sales:,.2f}</div>
            </div>
            <div class="summary-card">
                <div class="label">总订单</div>
                <div class="value orders">{int(global_orders)}</div>
            </div>
            <div class="summary-card">
                <div class="label">退货数量</div>
                <div class="value returns">{total_returns}</div>
            </div>
            <div class="summary-card">
                <div class="label">广告花费</div>
                <div class="value ads">${global_ads_spend:,.2f}</div>
            </div>
        </div>

        <!-- 图表区域 -->
        <div class="chart-grid">
            <div class="chart-card">
                <h3>📊 各国家销售额分布</h3>
                <div id="salesChart" class="chart"></div>
            </div>
            <div class="chart-card">
                <h3>🏆 Top 10 SKU 销售额</h3>
                <div id="skuChart" class="chart"></div>
            </div>
        </div>
        <div class="chart-grid">
            <div class="chart-card">
                <h3>📈 退货原因分布</h3>
                <div id="returnsReasonChart" class="chart"></div>
            </div>
            <div class="chart-card">
                <h3>🗺️ 各国家订单分布</h3>
                <div id="ordersChart" class="chart"></div>
            </div>
        </div>

        <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
        <script>
        // 图表数据
        const salesData = {sales_chart_data_str};
        const skuData = {sku_chart_data_str};
        const returnsReasonData = {returns_reason_data_str};
        const ordersData = {orders_chart_data_str};
        const countryNames = {country_names_list_str};

        // 初始化图表
        document.addEventListener('DOMContentLoaded', function() {{
            // 销售额饼图
            if (salesData.length > 0) {{
                const salesChart = echarts.init(document.getElementById('salesChart'));
                salesChart.setOption({{
                    tooltip: {{ trigger: 'item', formatter: function(params) {{ return params.name + ': $' + params.value.toFixed(2) + ' (' + params.percent + '%)'; }} }},
                    series: [{{
                        type: 'pie',
                        radius: ['40%', '70%'],
                        avoidLabelOverlap: false,
                        itemStyle: {{ borderRadius: 10, borderColor: '#fff', borderWidth: 2 }},
                        label: {{ show: true, formatter: '{{b}}' }},
                        data: salesData
                    }}]
                }});
            }}

            // SKU柱状图
            if (skuData.length > 0) {{
                const skuChart = echarts.init(document.getElementById('skuChart'));
                skuChart.setOption({{
                    tooltip: {{ trigger: 'axis', axisPointer: {{ type: 'shadow' }} }},
                    grid: {{ left: '12%', right: '4%', bottom: '3%', containLabel: true }},
                    xAxis: {{ type: 'value', axisLabel: {{ formatter: '${{value}}' }} }},
                    yAxis: {{ type: 'category', data: skuData.map(i => i.name), inverse: true }},
                    series: [{{
                        type: 'bar',
                        data: skuData.map(i => i.value),
                        itemStyle: {{ color: '#4CAF50' }}
                    }}]
                }});
            }}

            // 退货原因饼图
            if (returnsReasonData.length > 0) {{
                const returnsChart = echarts.init(document.getElementById('returnsReasonChart'));
                returnsChart.setOption({{
                    tooltip: {{ trigger: 'item' }},
                    series: [{{
                        type: 'pie',
                        radius: '70%',
                        data: returnsReasonData,
                        itemStyle: {{ borderRadius: 5, borderColor: '#fff', borderWidth: 1 }}
                    }}]
                }});
            }}

            // 订单分布饼图
            if (ordersData.length > 0) {{
                const ordersChart = echarts.init(document.getElementById('ordersChart'));
                ordersChart.setOption({{
                    tooltip: {{ trigger: 'item' }},
                    series: [{{
                        type: 'pie',
                        radius: ['40%', '70%'],
                        data: ordersData,
                        itemStyle: {{ borderRadius: 10, borderColor: '#fff', borderWidth: 2 }}
                    }}]
                }});
            }}
        }});
        </script>

        <div class="card">
            <div class="card-title">💰 核心数据 ({', '.join([COUNTRY_MAP[c]['name'] for c in target_countries])})</div>
            <table>
                <thead>
                    <tr>
                        {core_headers}
                    </tr>
                </thead>
                <tbody>{rows}</tbody>
            </table>
        </div>
        
        <div class="card">
            <div class="card-title">📦 SKU销售明细 (Top 50)</div>
            <div class="filter-bar">
                <label>国家:</label>
                <select id="filterCountry">
                    <option value="">全部</option>
                    {country_options}
                </select>
                <label>SKU:</label>
                <input type="text" id="filterSku" placeholder="输入SKU筛选...">
            </div>
            <table id="skuTable">
                <thead>
                    <tr>
                        <th data-sort="sku" style="text-align:left;min-width:180px">SKU</th>
                        {"<th data-sort=\"sessions\" class=\"num\">流量</th>" if has_business_data else ""}
                        <th data-sort="country">国家</th>
                        <th data-sort="qty" class="num">销量</th>
                        {"<th data-sort=\"conv\" class=\"num\">转化率</th>" if has_business_data else ""}
                        <th data-sort="sales" class="num">销售额(本地)</th>
                        <th data-sort="sales_usd" class="num">销售额(USD)</th>
                    </tr>
                </thead>
                <tbody>{sku_rows}</tbody>
                <tfoot>
                    <tr style="font-weight:bold;background:#f0f4ff;">
                        <td>合计</td>
                        {"<td class=\"num\" id=\"skuSessionsSum\">0</td>" if has_business_data else ""}
                        <td></td>
                        <td class="num" id="skuQtySum">0</td>
                        {"<td class=\"num\" id=\"skuConvAvg\">{avg_conv_rate:.2f}%</td>" if has_business_data else ""}
                        <td class="num">-</td>
                        <td class="num" id="skuSalesSum">$0.00</td>
                    </tr>
                </tfoot>
            </table>
        </div>
        
        {returns_html}
        
        <footer style="text-align:center;padding:12px;color:#888;font-size:11px;">
            Amazon Report Analyzer
        </footer>
    </div>
    <script>
    const DEFAULT_EUR_RATE = {eur_rate};
    const DEFAULT_GBP_RATE = {gbp_rate};
    
    (function() {{
        const table = document.getElementById('skuTable');
        const tbody = table.querySelector('tbody');
        const rows = Array.from(tbody.querySelectorAll('tr'));
        const filterCountry = document.getElementById('filterCountry');
        const filterSku = document.getElementById('filterSku');
        let sortCol = 'sales', sortAsc = false;
        
        function filterAndSort() {{
            const country = filterCountry.value.toLowerCase();
            const sku = filterSku.value.toLowerCase();
            
            let filtered = rows.filter(row => {{
                const cells = row.querySelectorAll('td');
                const hasSessions = row.dataset.hasSessions === 'true';
                const idxCountry = hasSessions ? 2 : 1;
                const rowCountry = cells[idxCountry].textContent.toLowerCase();
                const rowSku = cells[0].textContent.toLowerCase();
                return (country === '' || rowCountry.includes(country)) &&
                       (sku === '' || rowSku.includes(sku));
            }});
            
            filtered.sort((a, b) => {{
                const cellsA = a.querySelectorAll('td');
                const cellsB = b.querySelectorAll('td');
                const hasSessions = a.dataset.hasSessions === 'true';
                const idxSessions = hasSessions ? 1 : -1;
                const idxCountry = hasSessions ? 2 : 1;
                const idxQty = hasSessions ? 3 : 2;
                const idxConv = hasSessions ? 4 : -1;
                const idxSalesLocal = hasSessions ? 5 : 3;
                const idxSalesUsd = hasSessions ? 6 : 4;
                let valA, valB;
                if (sortCol === 'sku') {{ valA = cellsA[0].textContent; valB = cellsB[0].textContent; }}
                else if (sortCol === 'sessions') {{ 
                    if (idxSessions < 0) {{ valA = 0; valB = 0; }}
                    else {{ valA = parseInt(cellsA[idxSessions].textContent.replace(/,/g, '')) || 0; valB = parseInt(cellsB[idxSessions].textContent.replace(/,/g, '')) || 0; }}
                }}
                else if (sortCol === 'country') {{ valA = cellsA[idxCountry].textContent; valB = cellsB[idxCountry].textContent; }}
                else if (sortCol === 'qty') {{ valA = parseInt(cellsA[idxQty].textContent) || 0; valB = parseInt(cellsB[idxQty].textContent) || 0; }}
                else if (sortCol === 'conv') {{ 
                    if (idxConv < 0) {{ valA = 0; valB = 0; }}
                    else {{ valA = parseFloat(cellsA[idxConv].textContent.replace('%', '')) || 0; valB = parseFloat(cellsB[idxConv].textContent.replace('%', '')) || 0; }}
                }}
                else if (sortCol === 'sales_usd') {{ valA = parseFloat(cellsA[idxSalesUsd].getAttribute('data-usd')) || parseFloat(cellsA[idxSalesUsd].textContent.replace(/[$,€£]/g, '')) || 0; valB = parseFloat(cellsB[idxSalesUsd].getAttribute('data-usd')) || parseFloat(cellsB[idxSalesUsd].textContent.replace(/[$,€£]/g, '')) || 0; }}
                else {{ valA = parseFloat(cellsA[idxSalesLocal].textContent.replace(/[$,€£]/g, '')) || 0; valB = parseFloat(cellsB[idxSalesLocal].textContent.replace(/[$,€£]/g, '')) || 0; }}
                return sortAsc ? (valA > valB ? 1 : -1) : (valA < valB ? 1 : -1);
            }});
            
            tbody.innerHTML = '';
            filtered.forEach(row => tbody.appendChild(row));
            
            // 更新SKU求和
            let totalSessions = 0, totalQty = 0, totalSales = 0, weightedConv = 0;
            filtered.forEach(row => {{
                const cells = row.querySelectorAll('td');
                const hasSessions = row.dataset.hasSessions === 'true';
                const idxSessions = hasSessions ? 1 : -1;
                const idxQty = hasSessions ? 3 : 2;
                const idxConv = hasSessions ? 4 : -1;
                const idxSalesUsd = hasSessions ? 6 : 4;
                const sessions = idxSessions >= 0 ? parseInt(cells[idxSessions].textContent.replace(/,/g, '')) || 0 : 0;
                const qty = parseInt(cells[idxQty].textContent) || 0;
                const conv = idxConv >= 0 ? parseFloat(cells[idxConv].textContent.replace('%', '')) || 0 : 0;
                const sales = parseFloat(cells[idxSalesUsd].getAttribute('data-usd')) || parseFloat(cells[idxSalesUsd].textContent.replace(/[$,]/g, '')) || 0;
                totalSessions += sessions;
                totalQty += qty;
                totalSales += sales;
                weightedConv += sessions * conv;
            }});
            const avgConv = totalSessions > 0 ? (weightedConv / totalSessions).toFixed(2) + '%' : '-';
            document.getElementById('skuSessionsSum').textContent = totalSessions.toLocaleString();
            document.getElementById('skuQtySum').textContent = totalQty.toLocaleString();
            document.getElementById('skuConvAvg').textContent = avgConv;
            document.getElementById('skuSalesSum').textContent = '$' + totalSales.toFixed(2).replace(/\\B(?=(\\d{3})+(?!\\d))/g, ',');
        }}
        
        filterCountry.addEventListener('change', filterAndSort);
        filterSku.addEventListener('input', filterAndSort);
        
        table.querySelectorAll('th[data-sort]').forEach(th => {{
            th.addEventListener('click', () => {{
                const col = th.dataset.sort;
                if (sortCol === col) sortAsc = !sortAsc;
                else {{ sortCol = col; sortAsc = false; }}
                table.querySelectorAll('th').forEach(t => t.className = '');
                th.className = sortAsc ? 'sort-asc' : 'sort-desc';
                filterAndSort();
            }});
        }});
        
        filterAndSort();
    }})();
    
    function recalcRates() {{
        const rateEur = parseFloat(document.getElementById('rateEurUsd').value) || DEFAULT_EUR_RATE;
        const rateGbp = parseFloat(document.getElementById('rateGbpUsd').value) || DEFAULT_GBP_RATE;
        
        document.querySelectorAll('#skuTable tbody tr').forEach(row => {{
            const cells = row.querySelectorAll('td');
            const hasSessions = row.dataset.hasSessions === 'true';
            const idxSalesUsd = hasSessions ? 6 : 4;
            const currency = row.dataset.currency;
            const isEur = currency === 'EUR';
            const rate = isEur ? rateEur : rateGbp;
            const defaultRate = isEur ? DEFAULT_EUR_RATE : DEFAULT_GBP_RATE;

            let origUsd = parseFloat(cells[idxSalesUsd].getAttribute('data-usd'));
            if (isNaN(origUsd) || origUsd === 0) {{
                origUsd = parseFloat(cells[idxSalesUsd].textContent.replace(/[$,]/g, '')) || 0;
                cells[idxSalesUsd].setAttribute('data-usd', origUsd);
            }}
            const newUsd = origUsd * rate / defaultRate;
            cells[idxSalesUsd].textContent = '$' + newUsd.toFixed(2).replace(/\\B(?=(\\d{3})+(?!\\d))/g, ',');
        }});
        
        document.querySelectorAll('.card table tbody tr').forEach(row => {{
            const cells = row.querySelectorAll('td');
            cells.forEach(cell => {{
                const text = cell.textContent;
                if ((text.includes('€') || text.includes('£')) && text.includes('$')) {{
                    let origUsd = parseFloat(cell.getAttribute('data-usd'));
                    if (isNaN(origUsd)) {{
                        const match = text.match(/\\$(\\d[\\d,]*\\.\\d+)/);
                        if (match) {{
                            origUsd = parseFloat(match[1].replace(/,/g, ''));
                            cell.setAttribute('data-usd', origUsd);
                        }}
                    }}
                    if (!isNaN(origUsd)) {{
                        const newUsd = origUsd * rateEur / DEFAULT_EUR_RATE;
                        cell.textContent = text.replace(/\\$(\\d[\\d,]*\\.\\d+)/, '$' + newUsd.toFixed(2).replace(/\\B(?=(\\d{3})+(?!\\d))/g, ','));
                    }}
                }}
            }});
        }});
    }}
    
    // 初始化退货产品合计
    (function() {{
        const returnsTable = document.getElementById('returnsTable');
        if (returnsTable) {{
            const returnsRows = returnsTable.querySelectorAll('tbody tr');
            let totalReturns = 0;
            returnsRows.forEach(row => {{
                const cells = row.querySelectorAll('td');
                if (cells.length >= 2) {{
                    totalReturns += parseInt(cells[1].textContent) || 0;
                }}
            }});
            document.getElementById('returnsQtySum').textContent = totalReturns;
        }}
        
        const reasonTable = document.getElementById('reasonTable');
        if (reasonTable) {{
            const reasonRows = reasonTable.querySelectorAll('tbody tr');
            let totalReasons = 0;
            reasonRows.forEach(row => {{
                const cells = row.querySelectorAll('td');
                if (cells.length >= 2) {{
                    totalReasons += parseInt(cells[1].textContent) || 0;
                }}
            }});
            document.getElementById('reasonQtySum').textContent = totalReasons;
        }}
    }})();
    </script>
</body>
</html>"""
    
    if not output_file:
        # 使用安全的文件名
        safe_period = sanitize_filename(period)
        if not safe_period:
            safe_period = "report"
        output_file = f"amazon_report_{safe_period}.html"
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"✅ 报告已生成: {output_file}")
    print(f"   国家: {', '.join([COUNTRY_MAP[c]['name'] for c in target_countries])}")
    print(f"   周期: {period}")

    # 导出数据（如果指定了导出格式）
    if export_format:
        export_data = prepare_export_data(country_data, ads_data, returns, rates)

        if not export_filename:
            safe_period = sanitize_filename(period)
            if not safe_period:
                safe_period = "report"
            export_filename = f"amazon_report_{safe_period}"

        if export_format == 'xlsx' and not export_filename.endswith('.xlsx'):
            export_filename += '.xlsx'
        elif export_format == 'csv' and not export_filename.endswith('.csv'):
            export_filename += '.csv'

        if export_format == 'csv':
            export_to_csv(export_data, export_filename)
        elif export_format == 'xlsx':
            export_to_excel(export_data, export_filename, rates)

    return output_file


def run_interactive_mode(args):
    """交互式运行模式"""
    import sys
    print("\n" + "=" * 50)
    print("   Amazon Report Analyzer")
    print("   交互式模式")
    print("=" * 50)

    # 1. 询问目录
    default_dir = "amazon"
    if os.path.exists(default_dir):
        default_dir_hint = f" (直接回车使用: {default_dir})"
    else:
        default_dir_hint = ""

    dir_input = input(f"\n请输入数据报告目录路径{default_dir_hint}: ").strip()
    if not dir_input:
        dir_input = default_dir

    args.dir = dir_input

    # 2. 扫描目录获取可用国家
    validated_dir = validate_report_dir(args.dir)
    if validated_dir is None:
        print(f"错误: 目录不存在或无效 - {args.dir}")
        sys.exit(1)
    args.dir = validated_dir

    scan = scan_reports_directory(args.dir)
    available_countries = scan.get('countries', [])

    # 3. 询问国家
    if available_countries:
        country_names = []
        for c in available_countries:
            cname = COUNTRY_MAP.get(c, {}).get('name', c)
            country_names.append(f"{c} ({cname})")

        print(f"\n可用国家: {', '.join(country_names)}")
        countries_input = input("请输入要分析的国家代码，用空格分隔 (直接回车分析全部): ").strip()

        if countries_input:
            args.countries = countries_input.upper().split()
        else:
            args.countries = None
    else:
        args.countries = None
        print("\n未在目录中发现国家数据")

    # 4. 询问周期
    detected_period = scan.get('period', '')
    if detected_period:
        period_yyyymm = detected_period.replace('年', '').replace('月', '').replace('Q', '').replace(' ', '')
        if len(period_yyyymm) >= 6:
            period_yyyymm = period_yyyymm[:6]
        period_hint = f" (直接回车使用: {period_yyyymm})"
    else:
        period_hint = ""

    period_input = input(f'\n请输入报告周期 (yyyymm格式，如 202602){period_hint}: ').strip()
    if not period_input:
        period_input = detected_period

    args.period = period_input if period_input else None

    # 5. 询问是否导出数据
    print("\n导出选项:")
    print("  1 - 不导出")
    print("  2 - 导出 CSV")
    print("  3 - 导出 XLSX")
    export_choice = input("请选择 (直接回车默认不导出): ").strip()

    if export_choice == '2':
        args.export = 'csv'
    elif export_choice == '3':
        args.export = 'xlsx'
    else:
        args.export = None

    # 6. 询问导出文件名（如果需要导出）
    if args.export:
        export_hint = " (直接回车自动生成)"
        export_input = input(f"\n请输入导出文件名{export_hint}: ").strip()
        args.export_file = export_input if export_input else None
    else:
        args.export_file = None

    # 7. 询问汇率（仅导出XLSX时）
    args.eur_rate = None
    args.gbp_rate = None
    if args.export == 'xlsx':
        print("\n汇率设置 (直接回车使用默认值):")
        eur_input = input("  EUR→USD 汇率 (默认 1.0800): ").strip()
        gbp_input = input("  GBP→USD 汇率 (默认 1.2700): ").strip()
        if eur_input:
            try:
                args.eur_rate = float(eur_input)
            except ValueError:
                pass
        if gbp_input:
            try:
                args.gbp_rate = float(gbp_input)
            except ValueError:
                pass

    # 8. 询问输出文件名
    output_hint = " (直接回车自动生成)"
    output_input = input(f"\n请输入HTML报告文件名{output_hint}: ").strip()
    args.output = output_input if output_input else None

    print("\n" + "-" * 50)
    print("正在分析...")
    print("-" * 50 + "\n")

    return args


# ==================== 主程序 ====================
def find_report_dirs(base_dir=None):
    """自动查找当前目录及父目录下的报告目录"""
    original_dir = os.getcwd()
    
    if base_dir and os.path.exists(base_dir):
        os.chdir(base_dir)
    
    current_dir = os.getcwd()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    report_keywords = ['report', 'Report', '2025', '2026', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    dirs = []
    
    def is_report_dir(dir_path):
        """检查目录是否为报告目录"""
        if not os.path.isdir(dir_path):
            return False
        try:
            csv_files = [f for f in os.listdir(dir_path) if f.endswith('.csv')]
            if not csv_files:
                return False
            # 检查是否包含报告相关的CSV文件
            for f in csv_files:
                f_lower = f.lower()
                if any(kw in f for kw in report_keywords):
                    return True
            return False
        except OSError:
            return False
    
    def scan_dirs(base_path, rel_base, priority):
        if not os.path.exists(base_path):
            return
        try:
            for item in os.listdir(base_path):
                path = os.path.join(base_path, item)
                if is_report_dir(path):
                    rel_path = os.path.relpath(path, rel_base)
                    dirs.append((rel_path, priority))
        except OSError:
            pass
    
    # 1. 扫描当前目录本身（最高优先级）
    if is_report_dir(current_dir):
        dirs.append(('.', 200))
    
    # 2. 扫描当前目录的子目录
    scan_dirs(current_dir, current_dir, priority=150)
    
    # 3. 扫描当前目录的 amazon 子目录
    amazon_path = os.path.join(current_dir, 'amazon')
    if is_report_dir(amazon_path):
        dirs.append(('amazon', 150))
    
    # 4. 扫描脚本目录
    scan_dirs(script_dir, current_dir, priority=50)
    
    # 5. 扫描脚本目录的 amazon
    script_amazon = os.path.join(script_dir, 'amazon')
    if is_report_dir(script_amazon):
        dirs.append(('amazon', 50))
    
    # 6. 扫描父目录
    parent_dir = os.path.dirname(current_dir)
    if parent_dir != current_dir:
        scan_dirs(parent_dir, current_dir, priority=10)
        parent_amazon = os.path.join(parent_dir, 'amazon')
        if is_report_dir(parent_amazon):
            dirs.append(('amazon', 10))
    
    # 去重，按优先级和时间排序
    # 名字叫 "amazon" 的目录额外增加优先级
    unique_dirs = {}
    for d, priority in dirs:
        path = os.path.join(current_dir, d) if d != '.' else current_dir
        if os.path.exists(path):
            try:
                mtime = os.path.getmtime(path)
                # 名字精确匹配 "amazon" 的目录额外 +100 优先级
                name_priority = 100 if d == 'amazon' else 0
                total_priority = priority + name_priority
                if d not in unique_dirs or total_priority > unique_dirs[d][0]:
                    unique_dirs[d] = (total_priority, mtime)
            except OSError:
                pass
    
    dirs_with_time = [(d, v[1], v[0]) for d, v in unique_dirs.items()]
    dirs_with_time.sort(key=lambda x: (x[2], x[1]), reverse=True)
    result = [d[0] for d in dirs_with_time]
    
    # 恢复原始目录
    if base_dir:
        os.chdir(original_dir)
    
    return result

def validate_report_dir(dir_path):
    """验证并规范化报告目录路径，防止目录遍历攻击"""
    if not dir_path:
        return None

    # 获取脚本所在目录作为基准目录
    script_dir = os.path.realpath(os.path.dirname(os.path.abspath(__file__)))
    current_dir = os.path.realpath(os.getcwd())

    # 解析目标路径
    if os.path.isabs(dir_path):
        target_dir = os.path.realpath(dir_path)
    else:
        target_dir = os.path.realpath(os.path.join(current_dir, dir_path))

    # 检查路径是否包含 ..
    if '..' in dir_path:
        print(f"警告: 路径不允许包含 '..' : {dir_path}")
        return None

    # 限制在当前目录及脚本目录的父目录范围内
    allowed_roots = [
        os.path.realpath(current_dir),
        os.path.realpath(script_dir),
        os.path.realpath(os.path.join(script_dir, '..')),
        os.path.realpath(os.path.join(current_dir, '..')),
    ]

    # 检查目标路径是否在允许范围内
    is_allowed = any(target_dir.startswith(root) for root in allowed_roots)

    if not is_allowed:
        print(f"警告: 路径超出允许范围 : {dir_path}")
        return None

    # 检查目录是否存在
    if not os.path.isdir(target_dir):
        print(f"警告: 目录不存在 : {dir_path}")
        return None

    return target_dir


def main():
    # 自动检测报告目录
    report_dirs = find_report_dirs()

    parser = argparse.ArgumentParser(description='Amazon Sales Report Analyzer')
    parser.add_argument('--dir', '-d', help='报告目录 (不指定则自动检测)')
    parser.add_argument('--countries', '-c', nargs='+', help='指定国家代码 (DE IT FR ES UK)')
    parser.add_argument('--output', '-o', help='输出文件名')
    parser.add_argument('--period', '-p', help='报告周期名称，如 "2026年2月"')
    parser.add_argument('--export', choices=['csv', 'xlsx'], help='导出格式: csv 或 xlsx')
    parser.add_argument('--export-file', help='导出文件名 (不指定则自动生成)')
    parser.add_argument('--eur-rate', type=float, help='EUR→USD 汇率 (如 1.08)')
    parser.add_argument('--gbp-rate', type=float, help='GBP→USD 汇率 (如 1.27)')
    parser.add_argument('--interactive', '-i', action='store_true', help='交互式模式')
    args = parser.parse_args()

    # 交互式模式
    if args.interactive or (not args.dir and not args.interactive and len(sys.argv) == 1):
        args = run_interactive_mode(args)

    # 验证并规范化目录参数
    if args.dir:
        validated_dir = validate_report_dir(args.dir)
        if validated_dir is None:
            print("错误: 无效的目录参数，将使用自动检测")
            args.dir = None
    
    print("=" * 50)
    print("Amazon Sales Report Analyzer")
    print("=" * 50)
    
    # 确定目录
    if args.dir:
        reports_dir = args.dir
    elif report_dirs:
        # 自动选择最新的报告目录
        reports_dir = report_dirs[0]
        if len(report_dirs) > 1:
            print(f"\n🔍 自动选择最新报告目录: {reports_dir}")
            print(f"   (其他目录: {', '.join(report_dirs[1:])})")
        else:
            print(f"\n🔍 自动检测到报告目录: {reports_dir}")
    else:
        if os.path.exists('amazon'):
            reports_dir = 'amazon'
        else:
            reports_dir = DEFAULT_REPORTS_DIR
    
    # 显示扫描信息
    scan = scan_reports_directory(reports_dir)
    print(f"\n📁 目录: {reports_dir}")
    print(f"📅 周期: {scan['period']}")
    print(f"🌍 可用国家: {', '.join([COUNTRY_MAP[c]['name'] for c in scan['countries']]) if scan['countries'] else '未找到'}")
    print(f"📄 文件类型: {', '.join([k for k,v in scan['files'].items() if v])}")
    
    # 确定周期
    period = args.period if args.period else scan['period']
    
    # 生成报告
    generate_report(reports_dir, args.output, args.countries, period, args.export, args.export_file, args.eur_rate, args.gbp_rate)
    
    print("\n" + "=" * 50)

if __name__ == "__main__":
    main()
